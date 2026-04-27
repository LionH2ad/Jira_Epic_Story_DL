import openpyxl
from constants import JiraFields
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def apply_excel_style(file_path, sheet_names):
    """엑셀 시트에 남색 헤더, 틀 고정, 열 너비 자동 조절 스타일을 적용합니다."""
    wb = openpyxl.load_workbook(file_path)
    
    header_fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for name in sheet_names:
        if name not in wb.sheetnames: continue
        ws = wb[name]
        ws.freeze_panes = "A2" # 1행 틀 고정

        # iter_cols()를 사용하여 열 단위로 안전하게 반복합니다.
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
            # col[0]은 해당 열의 헤더 셀입니다.
            col_letter = col[0].column_letter # 열 문자 (A, B, C...)
            header_value = col[0].value       # 1행에 적힌 헤더 이름 (Key, Issue id 등)

            # 1. 헤더 스타일 적용 (각 열의 첫 번째 셀)
            header_cell = col[0]
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = align_center
            header_cell.border = thin_border

            # 2. 데이터 셀 스타일 적용
            for cell in col:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = align_left

            # 3. [핵심] 헤더 이름을 기준으로 너비 결정
            # COLUMN_WIDTHS 딕셔너리에 이름이 있으면 그 값을 쓰고, 없으면 기본값 사용
            target_width = JiraFields.COLUMN_WIDTHS.get(header_value, JiraFields.DEFAULT_WIDTH)
            ws.column_dimensions[col_letter].width = target_width

    wb.save(file_path)
