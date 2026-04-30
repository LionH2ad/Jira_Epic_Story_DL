import openpyxl
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from common.config import JiraConfig
from common.constants import JiraFields
from services.parser import parse_issue_info
from services.excel_style import apply_excel_style

# 3. 메인 저장 프로세스
def process_and_save(raw_response):
    if not raw_response or "issues" not in raw_response:
        print("가공할 데이터가 없습니다.")
        return
    
    # [1. D 드라이브 경로 및 서비스명 설정]
    # 상위 폴더명(Nissan_DB_DL_ESR)을 서비스 이름으로 추출
    service_name = os.path.basename(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    # 최종 저장 경로: D:/NISSAN_JIRA_DATA/Nissan_DB_DL_ESR
    excel_dir = JiraConfig.get_excel_path(service_name)
    
    try:
        os.makedirs(excel_dir, exist_ok=True) # 폴더 없으면 자동으로 생성
    except Exception as e:
        print(f"❌ D 드라이브 경로를 생성할 수 없습니다. 권한을 확인하세요: {e}")
        # 대안으로 현재 서비스 폴더 내 Excel 폴더 사용
        service_root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_dir = os.path.join(service_root_dir, "Excel")
        os.makedirs(excel_dir, exist_ok=True)
        print(f"⚠️ 대안 경로로 변경됨: {excel_dir}")

    # [2. 파일명 설정]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    # 파일명에도 서비스 이름을 포함하여 구분하기 쉽게 설정
    excel_file_path = os.path.join(excel_dir, f"DB_{service_name}_{timestamp}.xlsx")
   
    # 탭별 조합(Combination) 리스트
    dbc_tab_list = [] # Sheet 1: CAN DBC
    aasp_tab_list = [] # Sheet 1: aasp

    active_index = 1 # Status가 Close가 아닌 티켓용 번호
    active_aasp_index = 1
    base_url = "https://spaws.jp.nissan.biz/jira/browse/"

    for issue in raw_response["issues"]:
        # 1단계: 모든 정보가 포함된 info 데이터 추출
        info = parse_issue_info(issue) # parse_issue_info
        is_lge = str(info["Supplier"]).strip() == JiraFields.TARGET_SUPPLIER
        is_epic = info["Type"] == "Epic"
        
        # 2단계 
        # Sheet 1 CAN DBC
        # 조건: 첨부파일 리스트에 .dbc 파일이 하나라도 있는 경우
        if info["DBC_Files"] and is_lge and is_epic:
            status = info["Status"]
            key = info["Key"]
            
            # Index 결정: Status가 Close(또는 CLOSED)가 아니면 번호 증가, 맞으면 빈칸
            if status.upper() not in ["CLOSE", "CLOSED"]:
                dbc_idx_val = active_index
                active_index += 1
            else:
                dbc_idx_val = " " # 빈칸 (공백 한 칸)

            # .dbc 파일이 여러 개일 경우를 대비해 공백으로 합침
            dbc_name_str = "\n".join(info["DBC_Files"])

            dbc_tab_list.append({
                "Index": dbc_idx_val,
                "Ticket": key,
                "Link": key,
                "URL": f"{base_url}{key}",
                "Supplier": info["Supplier"],
                "Type": info["Type"],
                "Status": status,
                "Domain": info["Domain"],
                "Summary": info["Summary"],
                "Assignee": info["Assignee"],
                "Assignee Email": info["Assignee_Email"],
                "DBC File Name": dbc_name_str
            })

        # Sheet 2 SWD
        # 조건: SWD_Files 리스트가 비어있지 않은 모든 티켓
        if info["AASP_Combined_Items"]:
            status = info["Status"]
            key = info["Key"]
            
            # Index 로직
            if status.upper() not in ["CLOSE", "CLOSED"]:
                swd_idx_val = active_aasp_index
                active_aasp_index += 1
            else:
                swd_idx_val = " "

            aasp_tab_list.append({
                "Index": swd_idx_val,
                "Ticket": key,
                "Link": key,      # 나중에 하이퍼링크 처리
                "URL": f"{base_url}{key}",
                "Supplier": info["Supplier"],
                "Type": info["Type"],
                "Status": status,
                "Domain": info["Domain"],
                "Summary": info["Summary"],
                "Assignee": info["Assignee"],
                "Assignee Email": info["Assignee_Email"],
                "File Name": "\n".join(info["AASP_Combined_Items"]) # 엔터 구분
            })

    # 3단계: 멀티 탭 엑셀 저장
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        # 데이터가 있는 경우에만 시트 생성
        if dbc_tab_list:
            pd.DataFrame(dbc_tab_list).to_excel(writer, sheet_name=JiraFields.SHEET_DBC, index=False)
        if aasp_tab_list:
            pd.DataFrame(aasp_tab_list).to_excel(writer, sheet_name='SWD_Check_List', index=False)

    # 2. [핵심] openpyxl로 파일을 다시 열어 수식을 직접 입력
    wb = load_workbook(excel_file_path)
    for s_name in ['DBC_File_List', 'SWD_Check_List']:
        if s_name in wb.sheetnames:
            ws = wb[s_name]
            # 2행부터 데이터 끝까지 반복
            for row in range(2, ws.max_row + 1):
                key_cell = ws[f"B{row}"] # Ticket (B열)
                link_cell = ws[f"C{row}"] # Link (C열)
                url_text = ws[f"D{row}"].value # 이미 저장된 URL 주소 (D열)

                if key_cell.value:
                    link_cell.hyperlink = url_text # 링크처럼 보이게 스타일 설정 (파란색 + 밑줄)
                    link_cell.font = openpyxl.styles.Font(color="0563C1", underline="single")

    wb.save(excel_file_path)

    # 4단계: 디자인 스타일 적용 (생성된 모든 시트 대상)
    sheet_names = writer.sheets.keys() # 실제로 생성된 시트 이름만 가져오기
    apply_excel_style(excel_file_path, list(sheet_names))
    
    print(f"성공! 멀티 탭 파일 생성됨: {excel_file_path}")
