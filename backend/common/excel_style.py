import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def apply_excel_style(file_path, sheet_names, theme):
    wb = openpyxl.load_workbook(file_path)
    styles = theme["styles"]
    rules = theme["rules"]

    # 재사용 스타일 객체 생성
    header_fill = PatternFill(start_color=styles["header_bg"], end_color=styles["header_bg"], fill_type="solid")
    header_font = Font(color=styles["header_font"], bold=True, name="맑은 고딕")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for name in sheet_names:
        if name not in wb.sheetnames: continue
        ws = wb[name]
        ws.freeze_panes = "A2"

        for col in ws.iter_cols(min_row=1):
            header_cell = col[0]
            header_name = str(header_cell.value).strip() if header_cell.value else "default"
            col_letter = header_cell.column_letter
            
            # 명세서(rules)에서 해당 컬럼의 룰 획득
            rule = rules.get(header_name, rules["default"])
            ws.column_dimensions[col_letter].width = rule["width"]

            for cell in col:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill, cell.font = header_fill, header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=rule["wrap"])
    
    wb.save(file_path)